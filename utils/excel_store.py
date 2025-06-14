import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

DATA_DIR = "data/active_games"
TEMPLATE_PATH = "data/templates/Game_Template.xlsx"

class ExcelGameStore:
    def __init__(self, game_id):
        self.game_id = str(game_id)
        self.file_path = os.path.join(DATA_DIR, f"{self.game_id}.xlsx")

    def create_game_file(self, players, starting_cash, max_trades, start_date, end_date):
        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError("Template not found at expected path.")

        os.makedirs(DATA_DIR, exist_ok=True)
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb["GameInfo"]

        ws["B1"].value = self.game_id
        ws["B2"].value = start_date
        ws["B3"].value = end_date
        ws["B4"].value = starting_cash
        ws["B5"].value = max_trades
        ws["B6"].value = ", ".join(players)

        wb.save(self.file_path)

        leaderboard_df = pd.DataFrame({
        "Player": players,
        "Cash": [starting_cash] * len(players),
        "PortfolioValue": [0.0] * len(players),
        "NetWorth": [starting_cash] * len(players)
        })
        self.write_sheet("Leaderboard", leaderboard_df)

        holdings_df = pd.DataFrame(columns=["Player", "StockSymbol", "Shares", "CurrentPrice", "TotalValue"])
        self.write_sheet("PlayerHoldings", holdings_df)

        tradequeue_df = pd.DataFrame(columns=["TradeID", "Player", "StockSymbol", "Action", "Shares", "RequestedAt", "Status", "Notes"])
        self.write_sheet("TradeQueue", tradequeue_df)

        transactions_df = pd.DataFrame(columns=["TransactionID", "Player", "StockSymbol", "Action", "Shares", "Price", "TotalValue", "ExecutedAt"])
        self.write_sheet("Transactions", transactions_df)

    def load_game_info(self):
        df = pd.read_excel(self.file_path, sheet_name="GameInfo", header=None, index_col=0)

        def safe_date(val, fallback=None):
            try:
                return pd.to_datetime(val)
            except Exception:
                return fallback if fallback else datetime.today()

        def safe_float(val, default=0.0):
            try:
                return float(str(val).strip())
            except Exception:
                return default

        def safe_int(val, default=0):
            try:
                return int(float(str(val).strip()))
            except Exception:
                return default

        def safe_str(val, default=""):
            try:
                return str(val).strip()
            except Exception:
                return default

        def safe_list(val, default=[]):
            try:
                return [p.strip() for p in str(val).split(",") if p.strip()]
            except Exception:
                return default

        def get_cell(df, label, default=None):
            try:
                return df.at[label, 1]
            except Exception:
                return default

        info = {
            "GameID": safe_str(get_cell(df, "Game ID"), "UNKNOWN"),
            "StartDate": safe_date(get_cell(df, "Start Date"), datetime.today()),
            "EndDate": safe_date(get_cell(df, "End Date"), datetime.today()),
            "StartingCash": safe_float(get_cell(df, "Starting Cash"), 1000),
            "MaxTradesPerDay": safe_int(get_cell(df, "Max Trades Per Day"), 3),
            "Players": safe_list(get_cell(df, "Players"), [])
        }

        if pd.isna(info["StartDate"]) or pd.isna(info["EndDate"]):
            raise ValueError("Start Date or End Date is missing or invalid in GameInfo sheet.")

        return info

    def read_sheet(self, sheet_name):
        return pd.read_excel(self.file_path, sheet_name=sheet_name)

    def write_sheet(self, sheet_name, df):
        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    def append_trades(self, trade_df):
        existing = self.read_sheet("TradeQueue")
        updated = pd.concat([existing, trade_df], ignore_index=True)
        self.write_sheet("TradeQueue", updated)

    def update_holdings(self, new_holdings_df):
        self.write_sheet("PlayerHoldings", new_holdings_df)

    def update_leaderboard(self, leaderboard_df):
        self.write_sheet("Leaderboard", leaderboard_df)

    def get_transaction_history(self):
        return self.read_sheet("Transactions")

    def log_transaction(self, transaction_row):
        existing = self.read_sheet("Transactions")
        updated = pd.concat([existing, pd.DataFrame([transaction_row])], ignore_index=True)
        self.write_sheet("Transactions", updated)

    def game_exists(self):
        return os.path.exists(self.file_path)

    def get_path(self):
        return self.file_path
